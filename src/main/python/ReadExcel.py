# -*- encoding:utf-8 -*-
from openpyxl import load_workbook
import os

path = "D:\InterFaceTestCase.xlsx"
workbook = load_workbook(path)
print(workbook.sheetnames)
sheet = workbook.get_sheet_by_name("测试数据")
print(sheet["I"])   #显示存在数据的行  第I列
print(sheet[4]) #显示存在数据的列 第4行
data = sheet["B1"]
print(data.value) #获取B4的值
print("最大行：",sheet.max_row)
print("最大列：",sheet.max_column)
for n in sheet["D"]:
    print(n.value,end="\n")

